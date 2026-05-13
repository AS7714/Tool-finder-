"""
JOA AI Disclosure Scanner
=========================
Scans all PDF papers across ScienceDirect subfolders,
extracts Article History dates and AI declaration status,
and writes results to a formatted Excel file.

Usage:
    python joa_ai_scanner.py

Requirements:
    pip install pdfplumber openpyxl rapidfuzz
"""

import os
import re
import sys
from pathlib import Path
from datetime import datetime

try:
    import pdfplumber
except ImportError:
    sys.exit("ERROR: pdfplumber not installed. Run: pip install pdfplumber openpyxl rapidfuzz")

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install pdfplumber openpyxl rapidfuzz")

try:
    from rapidfuzz import fuzz
except ImportError:
    sys.exit("ERROR: rapidfuzz not installed. Run: pip install pdfplumber openpyxl rapidfuzz")


# ─────────────────────────────────────────────
# CONFIGURATION — edit ROOT_FOLDER if needed
# ─────────────────────────────────────────────
ROOT_FOLDER = r""
OUTPUT_FILE = os.path.join(ROOT_FOLDER, "AI_Disclosure_Results.xlsx")

# ─────────────────────────────────────────────
# KEYWORD TIERS FOR AI DETECTION
#
# Designed around the real Elsevier/JOA template:
#   Section header: "Declaration of Generative AI and
#                    AI-Assisted Technologies in the Writing Process"
#   Body pattern:   "During the preparation of this work, the author
#                    used [TOOL] in order to [REASON]. After using
#                    this tool/service, the author(s) reviewed and
#                    edited the content as needed..."
# ─────────────────────────────────────────────

# ── Tier 1: Formal section headers (exact + common variants) ──────────────
# These are the standardized bold section titles placed before references.
# We check tail of document first (where Elsevier mandates placement).
TIER1_HEADERS = [
    # The canonical Elsevier header (exact, as seen in image)
    "declaration of generative ai and ai-assisted technologies in the writing process",
    # Shortened variants authors actually use
    "declaration of generative ai",
    "declaration of ai-assisted technologies",
    "declaration of artificial intelligence",
    "generative ai and ai-assisted technologies",
    "ai-assisted technologies in the writing process",
    # Alternate phrasings seen in ortho/surgical literature
    "use of artificial intelligence in manuscript preparation",
    "use of ai in the writing process",
    "use of generative ai",
    "ai tools disclosure",
    "disclosure of ai use",
    "disclosure of generative ai",
    "role of ai in manuscript preparation",
    "ai writing assistance disclosure",
]

# ── Tier 2: High-confidence usage sentence anchors ───────────────────────
# Anchored to the Elsevier template body sentence AND real-world variants.
# Each pattern requires BOTH the trigger phrase AND an AI signal nearby.
# This is the anti-false-positive layer.
TIER2_PATTERNS = [
    # THE canonical Elsevier trigger sentence (as seen in the image)
    r"during the preparation of this (work|manuscript|article|paper).{0,300}(used|employ|utiliz)",

    # "the author(s) used [TOOL] in order to / to [reason]"
    r"(the author|the authors|we|i)\s+(used|utilized|employed)\s+\w[\w\s\-\.]{0,40}"
    r"\s+(in order to|to)\s+(check|correct|improve|enhance|assist|draft|edit|refine|generate|translate|reword|rephrase|summarize|proofread|review)",

    # "After using this tool/service, the author(s) reviewed..."  — Elsevier closing sentence
    r"after using this (tool|service|software|model|system).{0,200}(reviewed|edited|revised|checked)",

    # "used [named tool] to [verb]" — direct attribution
    r"(used|utilized|employed)\s+(claude|chatgpt|gpt[\s\-]?\d*o?|copilot|gemini|bard|dall[\s\-]?e|"
    r"midjourney|grammarly|deepl|perplexity|llama|mistral|cohere|bing\s+ai|google\s+ai|"
    r"amazon\s+bedrock|anthropic|openai|grok)\b.{0,200}(to|for|in order)",

    # "AI was used to [verb]" — compact usage statement
    r"\b(generative\s+)?ai\s+(was|has been|were)\s+(used|utilized|employed|applied)\s+(to|for|in)",

    # "LLM / large language model was used"
    r"\b(large language model|llm)s?\s+(was|were|has been|have been)\s+(used|utilized|employed|applied)",

    # "AI-assisted [writing/editing/drafting...]"  — only in declaration context
    r"\bai[- ](assisted|generated|supported)\s+(writing|editing|drafting|language\s+editing|"
    r"manuscript\s+preparation|text\s+preparation|translation|grammar|proofreading)",

    # "with the assistance / help of [AI tool name]"
    r"with (the )?(assistance|help|aid|support) of\s+(claude|chatgpt|gpt|copilot|gemini|"
    r"bard|an?\s+(ai|llm|language\s+model|generative\s+ai))",

    # "takes full responsibility for the content" — Elsevier boilerplate closing line
    # Only meaningful if near a declaration section — checked only in tail text
    r"takes?\s+full\s+responsibilit(y|ies)\s+for\s+the\s+(content|publication|article)",
]

# ── Tier 3: Named AI tools (matched only inside declaration-like context) ─
# Matched as whole words. Checked after Tier 1/2 to avoid standalone
# mentions of "Claude" as an author name, etc.
AI_TOOLS = [
    # OpenAI
    "chatgpt", "gpt-4o", "gpt-4", "gpt-3.5", "gpt-3", "gpt4o", "gpt4", "gpt3", "openai",
    # Anthropic
    "claude 3.5", "claude 3", "claude sonnet", "claude opus", "claude haiku",
    "claude 2", "claude instant", "anthropic",
    # Google
    "gemini", "gemini pro", "gemini ultra", "bard", "google ai",
    # Microsoft
    "copilot", "bing ai", "bing chat", "microsoft ai",
    # Meta / Open source
    "llama", "llama 2", "llama 3", "mistral", "mixtral",
    # Image / multimodal
    "dall-e", "dalle", "midjourney", "stable diffusion", "sora",
    # Writing / productivity AI
    "grammarly", "deepl", "writesonic", "jasper", "copy.ai",
    # Research AI tools
    "perplexity", "elicit", "consensus", "research rabbit", "scite",
    "semantic scholar", "connected papers",
    # Other LLMs
    "cohere", "command r", "amazon bedrock", "grok", "inflection",
]

# ── Tier 4: Softer signals (only trigger if NOT already detected) ──────────
# These need broader context + lower confidence to avoid false positives.
# We require them to appear in the tail section of the doc (last 30%).
TIER4_PATTERNS = [
    # "AI was used" — generic
    r"\bai\s+was\s+used\b",
    r"\bartificial intelligence\s+was\s+(used|employed|utilized)\b",
    # "AI-generated content"
    r"\bai[- ]generated\s+(content|text|output|summary|response|draft)\b",
    # "language model" standalone
    r"\blanguage\s+model(s)?\b",
    # "machine-generated"
    r"\bmachine[- ]generated\s+(text|content|output|draft)\b",
    # "AI tool(s)" without further context
    r"\bai\s+tool(s)?\b",
    # "AI assistance" / "AI support"
    r"\bai\s+(assistance|support|help)\b",
    r"\bassisted\s+by\s+(an?\s+)?ai\b",
]

# ── Tier 5: NEGATIVE declarations (formally states NO AI was used) ─────────
# Placed in the same section header location — must be distinguished from
# positive declarations. Returned separately so Excel column is accurate.
TIER5_NEGATIVE = [
    r"no\s+(generative\s+)?ai\s+(tool|tools|technology|technologies)?\s*(was|were)\s+used",
    r"ai\s+(tool|tools|technology|technologies)?\s*(was|were)\s+not\s+used",
    r"did\s+not\s+use\s+(any\s+)?(generative\s+)?ai",
    r"no\s+ai[- ]assisted\s+(tools?|technologies?|writing|editing)",
    r"without\s+(the\s+)?use\s+of\s+(any\s+)?(generative\s+)?ai",
    r"ai\s+tools?\s+(have|has)\s+not\s+been\s+used",
    r"(the\s+)?(author|authors|we)\s+(did\s+not|have\s+not|has\s+not)\s+use[d]?\s+(any\s+)?"
    r"(generative\s+)?ai",
    r"no\s+large\s+language\s+model",
    r"not\s+use[d]?\s+any\s+(ai|artificial intelligence|llm|language\s+model)",
]

# ─────────────────────────────────────────────
# DATE EXTRACTION PATTERNS
# ─────────────────────────────────────────────
MONTH_NAMES = (
    "january|february|march|april|may|june|july|august|september|"
    "october|november|december|jan|feb|mar|apr|jun|jul|aug|sep|oct|nov|dec"
)

DATE_PATTERNS = {
    "received": [
        rf"received[:\s]+(\d{{1,2}}\s+(?:{MONTH_NAMES})\s+\d{{4}})",
        rf"received[:\s]+((?:{MONTH_NAMES})\s+\d{{1,2}},?\s+\d{{4}})",
        rf"received[:\s]+(\d{{1,2}}/\d{{1,2}}/\d{{4}})",
        rf"received[:\s]+(\d{{4}}-\d{{2}}-\d{{2}})",
    ],
    "revised": [
        rf"received in revised form[:\s]+(\d{{1,2}}\s+(?:{MONTH_NAMES})\s+\d{{4}})",
        rf"revised[:\s]+(\d{{1,2}}\s+(?:{MONTH_NAMES})\s+\d{{4}})",
        rf"revised manuscript received[:\s]+(\d{{1,2}}\s+(?:{MONTH_NAMES})\s+\d{{4}})",
        rf"revision received[:\s]+(\d{{1,2}}\s+(?:{MONTH_NAMES})\s+\d{{4}})",
    ],
    "accepted": [
        rf"accepted[:\s]+(\d{{1,2}}\s+(?:{MONTH_NAMES})\s+\d{{4}})",
        rf"accepted[:\s]+((?:{MONTH_NAMES})\s+\d{{1,2}},?\s+\d{{4}})",
        rf"accepted[:\s]+(\d{{1,2}}/\d{{1,2}}/\d{{4}})",
    ],
    "available_online": [
        rf"available online[:\s]+(\d{{1,2}}\s+(?:{MONTH_NAMES})\s+\d{{4}})",
        rf"published online[:\s]+(\d{{1,2}}\s+(?:{MONTH_NAMES})\s+\d{{4}})",
        rf"online[:\s]+(\d{{1,2}}\s+(?:{MONTH_NAMES})\s+\d{{4}})",
    ],
}


# ─────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────

def extract_text_from_pdf(pdf_path: Path) -> tuple[str, str]:
    """Extract full text and last-30%-of-doc text from a PDF."""
    full_text = ""
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            pages = pdf.pages
            total = len(pages)
            all_page_texts = []
            for page in pages:
                t = page.extract_text() or ""
                all_page_texts.append(t)
            full_text = "\n".join(all_page_texts)
            # Last 30% of pages for declaration (minimum last 3 pages)
            cutoff = max(0, total - max(3, int(total * 0.30)))
            tail_text = "\n".join(all_page_texts[cutoff:])
    except Exception as e:
        print(f"    [WARNING] Could not read {pdf_path.name}: {e}")
        return "", ""
    return full_text, tail_text


def extract_dates(text: str) -> dict:
    """Extract article history dates from text."""
    text_lower = text.lower()
    results = {}
    for field, patterns in DATE_PATTERNS.items():
        for pattern in patterns:
            match = re.search(pattern, text_lower)
            if match:
                raw = match.group(1).strip()
                results[field] = raw.title()
                break
        else:
            results[field] = ""
    return results


def detect_ai_declaration(full_text: str, tail_text: str) -> dict:
    """
    Detect AI declaration in the paper.

    Logic order (most → least specific to avoid false positives):
      1. Negative declarations  → "No (Null Declaration)"
      2. Tier 1 section header  → confirms a declaration block exists
      3. Tier 2 usage sentences → confirms AI was used + extracts evidence
      4. Tier 3 named AI tools  → only inside declaration window (~500 chars around header)
      5. Tier 4 soft signals    → only in tail text, lowest confidence
    """
    text_lower = full_text.lower()
    tail_lower = tail_text.lower()

    declared = False
    ai_used = "Not Detected"
    tools_found = []
    matched_text = ""
    confidence = "Low"
    section_found = ""
    declaration_window = ""  # Text near the declaration header for Tier 3 scoping

    # ── Tier 5: Negative declaration (search full text) ───────────────────
    # Must be checked in tail first (where it's placed), then full text.
    for pattern in TIER5_NEGATIVE:
        m = re.search(pattern, tail_lower) or re.search(pattern, text_lower)
        if m:
            # Only accept if near a declaration header OR in the tail
            context = _get_context(tail_lower if re.search(pattern, tail_lower) else text_lower,
                                   m.start(), m.end(), window=300)
            declared = True
            ai_used = "No (Null Declaration)"
            matched_text = context
            confidence = "High"
            section_found = "Negative Declaration"
            return _build_result(declared, ai_used, tools_found, matched_text, confidence, section_found)

    # ── Tier 1: Formal section header ─────────────────────────────────────
    # Search tail first (correct placement), fall back to full text.
    for header in TIER1_HEADERS:
        # Exact match
        for search_text in [tail_lower, text_lower]:
            if header in search_text:
                idx = search_text.find(header)
                declared = True
                section_found = header.title()
                confidence = "High"
                # Extract 600-char window AFTER the header for Tier 3 tool scoping
                declaration_window = search_text[idx: idx + 600]
                break

        if not declared:
            # Fuzzy match on individual lines (catches OCR errors, line breaks mid-header)
            for search_text in [tail_lower, text_lower]:
                for line in search_text.splitlines():
                    line = line.strip()
                    if len(line) < 15:
                        continue
                    score = fuzz.partial_ratio(header, line)
                    if score >= 88:
                        declared = True
                        section_found = f"{header.title()} (fuzzy match)"
                        confidence = "Medium"
                        idx = search_text.find(line)
                        declaration_window = search_text[idx: idx + 600]
                        break
                if declared:
                    break
        if declared:
            break

    # ── Tier 2: Usage sentence patterns ───────────────────────────────────
    # Search tail text first, then full text.
    # The last Tier 2 pattern (takes full responsibility) only runs on tail.
    for i, pattern in enumerate(TIER2_PATTERNS):
        search_scope = tail_lower if i == len(TIER2_PATTERNS) - 1 else text_lower
        m = re.search(pattern, search_scope, re.IGNORECASE | re.DOTALL)
        if m:
            declared = True
            ai_used = "Yes"
            matched_text = _get_context(search_scope, m.start(), m.end(), window=250)
            confidence = "High"
            if not section_found:
                section_found = "Inline Usage Statement"
            if not declaration_window:
                declaration_window = matched_text
            break

    # ── Tier 3: Named AI tools ─────────────────────────────────────────────
    # Only scan inside the declaration window (text around/after header) if we
    # already found a declaration. This prevents false positives from tool names
    # mentioned in the METHODS section (e.g., "AI-assisted surgery system").
    scan_for_tools = declaration_window if declaration_window else (tail_lower if declared else "")
    if scan_for_tools:
        for tool in AI_TOOLS:
            pattern = rf"(?<![a-z]){re.escape(tool)}(?![a-z])"
            if re.search(pattern, scan_for_tools, re.IGNORECASE):
                tools_found.append(tool.title())
                if ai_used == "Not Detected":
                    ai_used = "Yes"
                if confidence == "Low":
                    confidence = "Medium"

    # If we found tools but no Tier 2 sentence match, still set declared=True
    if tools_found:
        declared = True

    # ── Tier 4: Soft signals (tail text ONLY, no prior detection) ─────────
    # Only fires if nothing else has been confirmed yet. Lowest confidence.
    if ai_used == "Not Detected":
        for pattern in TIER4_PATTERNS:
            m = re.search(pattern, tail_lower, re.IGNORECASE)
            if m:
                declared = True
                ai_used = "Yes — Review Needed"
                matched_text = _get_context(tail_lower, m.start(), m.end(), window=200)
                confidence = "Low"
                if not section_found:
                    section_found = "Soft Signal (Tail)"
                break

    # ── Final state ────────────────────────────────────────────────────────
    # Section header found but no usage sentence or tool name found
    if declared and section_found and ai_used == "Not Detected":
        ai_used = "Review Needed — Section Found, No Tool Named"
        confidence = "Medium"

    # Deduplicate and title-case tools
    tools_found = list(dict.fromkeys(tools_found))

    return _build_result(declared, ai_used, tools_found, matched_text, confidence, section_found)


def _get_context(text: str, start: int, end: int, window: int = 200) -> str:
    """Get surrounding text context around a match."""
    ctx_start = max(0, start - window)
    ctx_end = min(len(text), end + window)
    snippet = text[ctx_start:ctx_end].replace("\n", " ").strip()
    return snippet[:300]  # Truncate for Excel cell


def _build_result(declared, ai_used, tools_found, matched_text, confidence, section_found):
    return {
        "declared": "Yes" if declared else "No",
        "ai_used": ai_used,
        "tools_found": ", ".join(tools_found) if tools_found else "",
        "matched_text": matched_text,
        "confidence": confidence,
        "section_found": section_found,
    }


def get_subfolders(root: Path) -> list[Path]:
    """Get subfolders to scan, sorted numerically when possible."""
    folders = []
    if not root.exists():
        return folders

    # Include all immediate child directories. The root path is already
    # ScienceDirect/year-scoped, and child folder names are often numeric.
    for item in root.iterdir():
        if item.is_dir():
            folders.append(item)

    # If PDFs are stored directly in root (no child folders), scan root itself.
    if not folders and any(root.glob("*.pdf")):
        return [root]

    # Sort by the first number in folder name
    def sort_key(p):
        nums = re.findall(r"\d+", p.name)
        return int(nums[0]) if nums else 0
    folders.sort(key=sort_key)
    return folders


# ─────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────

HEADERS = [
    "Paper Title",
    "Subfolder",
    "Received Date",
    "Revised Date",
    "Accepted Date",
    "Available Online",
    "AI Declaration Found?",
    "AI Used?",
    "AI Tools Mentioned",
    "Matched Text / Evidence",
    "Section Found",
    "Confidence",
]

COL_WIDTHS = [60, 22, 18, 18, 18, 18, 20, 22, 25, 60, 35, 12]

# Fill colors
FILL_YES    = PatternFill("solid", fgColor="C6EFCE")   # green
FILL_NO     = PatternFill("solid", fgColor="FFC7CE")   # red
FILL_REVIEW = PatternFill("solid", fgColor="FFEB9C")   # yellow
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")   # dark blue
FILL_ALT    = PatternFill("solid", fgColor="EBF3FB")   # light blue (alt row)

FONT_HEADER = Font(name="Arial", bold=True, color="FFFFFF", size=11)
FONT_BODY   = Font(name="Arial", size=10)
FONT_TITLE  = Font(name="Arial", bold=True, size=14, color="1F4E79")

BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def create_workbook(results: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active worksheet")
    ws.title = "AI Disclosure Results"

    # ── Title row ──
    ws.merge_cells("A1:L1")
    ws["A1"] = f"JOA AI Disclosure Scanner — Generated {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Summary row ──
    total = len(results)
    yes_count = sum(1 for r in results if "Yes" in r.get("ai_used", ""))
    no_count  = sum(1 for r in results if "No" in r.get("ai_used", ""))
    review_count = total - yes_count - no_count

    ws.merge_cells("A2:L2")
    ws["A2"] = (
        f"Total Papers: {total}   |   AI Used (Yes): {yes_count}   |   "
        f"No AI / Null Declaration: {no_count}   |   Review Needed / Not Detected: {review_count}"
    )
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Header row ──
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = BORDER_THIN
    ws.row_dimensions[3].height = 30

    # ── Data rows ──
    for row_idx, r in enumerate(results, start=4):
        ai_used = r.get("ai_used", "")
        is_alt = (row_idx % 2 == 0)

        row_data = [
            r.get("title", ""),
            r.get("subfolder", ""),
            r.get("received", ""),
            r.get("revised", ""),
            r.get("accepted", ""),
            r.get("available_online", ""),
            r.get("declared", ""),
            ai_used,
            r.get("tools_found", ""),
            r.get("matched_text", ""),
            r.get("section_found", ""),
            r.get("confidence", ""),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_BODY
            cell.border = BORDER_THIN

            # Alignment
            if col_idx in (1, 10, 11):
                cell.alignment = LEFT_WRAP
            else:
                cell.alignment = CENTER

            # Color coding based on AI Used column (col 8)
            if col_idx == 8:
                if "Yes" in ai_used and "Review" not in ai_used:
                    cell.fill = FILL_YES
                elif "No" in ai_used or "Null" in ai_used:
                    cell.fill = FILL_NO
                else:
                    cell.fill = FILL_REVIEW
            elif col_idx == 7:
                if r.get("declared") == "Yes":
                    cell.fill = FILL_YES
                else:
                    cell.fill = FILL_NO
            elif is_alt:
                cell.fill = FILL_ALT

        ws.row_dimensions[row_idx].height = 40

    # ── Column widths ──
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze panes at row 4 ──
    ws.freeze_panes = "A4"

    # ── Auto-filter on header row ──
    ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}3"

    # ── Summary sheet ──
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Category"
    ws2["B1"] = "Count"
    ws2["A1"].font = Font(bold=True, name="Arial")
    ws2["B1"].font = Font(bold=True, name="Arial")

    summary_data = [
        ("Total Papers Scanned", total),
        ("AI Used — Yes", yes_count),
        ("AI Not Used / Null Declaration", no_count),
        ("Review Needed / Not Detected", review_count),
        ("High Confidence Detections", sum(1 for r in results if r.get("confidence") == "High")),
        ("Medium Confidence Detections", sum(1 for r in results if r.get("confidence") == "Medium")),
        ("Low Confidence", sum(1 for r in results if r.get("confidence") == "Low")),
    ]
    for i, (label, count) in enumerate(summary_data, start=2):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=2, value=count).font = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 10

    return wb


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    root = Path(ROOT_FOLDER)

    if not root.exists():
        print(f"\n[ERROR] Root folder not found:\n  {ROOT_FOLDER}")
        print("  → Check the network path is accessible from your machine.")
        sys.exit(1)

    print(f"\n{'='*65}")
    print("  JOA AI Disclosure Scanner")
    print(f"{'='*65}")
    print(f"  Root folder : {ROOT_FOLDER}")
    print(f"  Output file : {OUTPUT_FILE}")
    print(f"{'='*65}\n")

    subfolders = get_subfolders(root)

    if not subfolders:
        print("[ERROR] No 'science direct' subfolders found in the root path.")
        sys.exit(1)

    print(f"  Found {len(subfolders)} subfolder(s).\n")

    results = []
    total_pdfs = 0
    errors = []

    for folder in subfolders:
        pdf_files = sorted(folder.glob("*.pdf"))
        if not pdf_files:
            print(f"  [SKIP] {folder.name} — no PDFs found")
            continue

        print(f"  [{folder.name}] — {len(pdf_files)} PDF(s)")

        for pdf_path in pdf_files:
            total_pdfs += 1
            title = pdf_path.stem  # filename without .pdf
            print(f"    → {title[:70]}{'...' if len(title) > 70 else ''}", end=" ")

            try:
                full_text, tail_text = extract_text_from_pdf(pdf_path)

                if not full_text.strip():
                    print("[EMPTY]")
                    errors.append(str(pdf_path))
                    results.append({
                        "title": title,
                        "subfolder": folder.name,
                        "received": "", "revised": "", "accepted": "", "available_online": "",
                        "declared": "Error", "ai_used": "Could Not Read PDF",
                        "tools_found": "", "matched_text": "", "section_found": "", "confidence": "",
                    })
                    continue

                dates = extract_dates(full_text)
                ai_result = detect_ai_declaration(full_text, tail_text)

                results.append({
                    "title": title,
                    "subfolder": folder.name,
                    **dates,
                    **ai_result,
                })

                status_icon = {
                    "Yes": "✓ AI Used",
                    "No (Null Declaration)": "✗ No AI",
                }.get(ai_result["ai_used"], "? Review")
                print(f"[{status_icon}]")

            except Exception as e:
                print(f"[ERROR: {e}]")
                errors.append(str(pdf_path))
                results.append({
                    "title": title,
                    "subfolder": folder.name,
                    "received": "", "revised": "", "accepted": "", "available_online": "",
                    "declared": "Error", "ai_used": f"Error: {str(e)[:80]}",
                    "tools_found": "", "matched_text": "", "section_found": "", "confidence": "",
                })

    print(f"\n{'='*65}")
    print(f"  Scan complete — {total_pdfs} PDFs processed across {len(subfolders)} folder(s)")
    if errors:
        print(f"  Errors (could not read): {len(errors)} file(s)")
        for e in errors[:5]:
            print(f"    - {e}")
        if len(errors) > 5:
            print(f"    ... and {len(errors)-5} more")
    print(f"{'='*65}\n")

    print("  Writing Excel output...")
    wb = create_workbook(results)
    wb.save(OUTPUT_FILE)
    print(f"  ✅ Saved: {OUTPUT_FILE}\n")


if __name__ == "__main__":
    main()
